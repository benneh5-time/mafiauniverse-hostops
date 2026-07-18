from __future__ import annotations

from unittest.mock import MagicMock

from openpyxl import load_workbook

from host_ops.commands import roles_export as rx


def test_extract_thread_id_from_various_urls():
    assert rx.extract_thread_id("https://www.mafiauniverse.com/forums/threads/60309-Some-Title") == "60309"
    assert rx.extract_thread_id("https://www.mafiauniverse.com/forums/threads/60309/") == "60309"
    assert rx.extract_thread_id("https://www.mafiauniverse.com/forums/showthread.php?t=60309") == "60309"
    assert rx.extract_thread_id("  60309  ") == "60309"
    assert rx.extract_thread_id("not a thread") is None
    assert rx.extract_thread_id("") is None


def test_unwrap_quote_strips_outer_wrapper_only():
    quote = "[QUOTE=MU Anniversary 2026;11042484][B]hi[/B]\n\n[QUOTE=Bob;1]nested[/QUOTE][/QUOTE]"
    body = rx.unwrap_quote_bbcode(quote)
    assert body == "[B]hi[/B]\n\n[QUOTE=Bob;1]nested[/QUOTE]"


def test_unwrap_quote_handles_surrounding_whitespace():
    quote = "\n  [QUOTE=Author;7]body[/QUOTE]\n\n"
    assert rx.unwrap_quote_bbcode(quote) == "body"


def test_unwrap_quote_without_wrapper_returns_trimmed():
    assert rx.unwrap_quote_bbcode("  plain body  ") == "plain body"


def test_safe_output_path_forces_xlsx_and_basename():
    assert rx._safe_output_path("roles.xlsx").name == "roles.xlsx"
    assert rx._safe_output_path("roles").name == "roles.xlsx"
    # directory traversal components are dropped
    assert rx._safe_output_path("../../etc/passwd").name == "passwd.xlsx"
    assert rx._safe_output_path("../../etc/passwd").parent == rx.EXPORT_DIR


def test_write_posts_xlsx_roundtrip(tmp_path):
    path = tmp_path / "out.xlsx"
    rows = [("Alice", "1", "[B]role a[/B]"), ("Bob", "2", "plain")]
    rx.write_posts_xlsx(path, rows)

    wb = load_workbook(path)
    ws = wb.active
    values = list(ws.iter_rows(values_only=True))
    assert values[0] == ("author", "post_id", "bbcode")
    assert values[1] == ("Alice", "1", "[B]role a[/B]")
    assert values[2] == ("Bob", "2", "plain")


def test_collect_rows_enumerates_and_unwraps(monkeypatch):
    monkeypatch.setattr(rx.time, "sleep", lambda *_: None)
    client = MagicMock()
    client.fetch_all_thread_posts.return_value = [
        {"post_id": 11, "author": "Alice"},
        {"post_id": 12, "author": "Bob"},
    ]
    client.fetch_quote_bbcode.side_effect = [
        "[QUOTE=Alice;11][B]role a[/B][/QUOTE]",
        "[QUOTE=Bob;12]role b[/QUOTE]",
    ]

    rows = rx._collect_rows(client, "60309")

    assert rows == [("Alice", "11", "[B]role a[/B]"), ("Bob", "12", "role b")]
    client.fetch_all_thread_posts.assert_called_once_with("60309")


def test_collect_rows_keeps_row_when_fetch_fails(monkeypatch):
    monkeypatch.setattr(rx.time, "sleep", lambda *_: None)
    client = MagicMock()
    client.fetch_all_thread_posts.return_value = [{"post_id": 11, "author": "Alice"}]
    client.fetch_quote_bbcode.side_effect = RuntimeError("boom")

    rows = rx._collect_rows(client, "60309")

    assert rows == [("Alice", "11", "")]


def test_collect_rows_falls_back_to_quote_author(monkeypatch):
    monkeypatch.setattr(rx.time, "sleep", lambda *_: None)
    client = MagicMock()
    client.fetch_all_thread_posts.return_value = [{"post_id": 11, "author": None}]
    client.fetch_quote_bbcode.return_value = "[QUOTE=QuotedName;11]body[/QUOTE]"

    rows = rx._collect_rows(client, "60309")

    assert rows == [("QuotedName", "11", "body")]
